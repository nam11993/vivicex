from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_sheet_as_dicts(file_path: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(header).strip() for header in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if all(value is None for value in row):
            continue
        records.append({headers[index]: value for index, value in enumerate(row)})
    return records
